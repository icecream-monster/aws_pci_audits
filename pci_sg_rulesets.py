"""
This script generates all rules per sg, per region, per account 
This script takes 3 commands: aws_account, sg_id, and region
"""
import boto3
import boto3.session
import sys
from datetime import datetime
import openpyxl
from openpyxl import Workbook, load_workbook

aws_account = sys.argv[1]
sg = sys.argv[2]
region = sys.argv[3]

my_session = boto3.session.Session(
    profile_name=aws_account,
    region_name=region)  # Establshing a session with your profile
ec2 = my_session.client("ec2")  # Using ec2 sdk
response = ec2.describe_security_groups(GroupIds=[sg])
sg_name = response["SecurityGroups"][0]["GroupName"]
# print(response["SecurityGroups"][0])

# Open a worksheet for current sg
workbook_name = datetime.now().strftime(
    "%Y-%m-%d") + "-pci_firewall_review.xlsx"
try:
    wb = load_workbook(filename=workbook_name)
except:
    wb = Workbook()
# New worksheet for
ws = wb.create_sheet(sg)
# Remove the annoying default named 'Sheet'
try:
    wb.remove(wb["Sheet"])
except:
    pass
# Worksheet headings
ws.append([aws_account, sg, sg_name, region])
ws.append(
    ["", "From port", "To Port", "Protocol", "Destination", "Description"])


# Func checking for all types of applicable connections
def check_sgs(egress_or_ingress):
    for rule in response["SecurityGroups"][0][egress_or_ingress]:
        # Check for ingress and egress rules
        if egress_or_ingress == "IpPermissionsEgress":
            rule_gress = "Egress"
        elif egress_or_ingress == "IpPermissions":
            rule_gress = "Ingress"
        # ports & protocols
        from_port = rule['FromPort'] if 'FromPort' in rule else "Any"
        to_port = rule['ToPort'] if "ToPort" in rule else "Any"
        protcol = rule["IpProtocol"]

        # Connections via PrefixListIds (e.g., s3 bucket, endpoints)
        for item in rule["PrefixListIds"]:
            if "PrefixListId" in item:
                dest = item["PrefixListId"]
                description = item["Description"]
            try:
                ws.append([
                    rule_gress, from_port, to_port, protcol, dest, description
                ])
            except:
                print("Rule failed to append to results: {}".format(rule))
                break

        # Connections via IPs directly, no VPC
        for item in rule["IpRanges"]:
            if "CidrIp" in item:
                dest = item["CidrIp"]
            if "Description" in item:
                description = item["Description"]
            else:
                description = "-"
            try:
                ws.append([
                    rule_gress, from_port, to_port, protcol, dest, description
                ])
            except:
                print("Rule failed to append to results: {}".format(rule))
                break

        # Connections to other SGs, likely via a VPC
        for item in rule["UserIdGroupPairs"]:
            dest = item["GroupId"]
            description = item["Description"]
            try:
                ws.append([
                    rule_gress, from_port, to_port, protcol, dest, description
                ])
            except:
                print("Rule failed to append to results: {}".format(rule))
                break
    wb.save(workbook_name)


check_sgs("IpPermissions")  # Check for ingress rules
check_sgs("IpPermissionsEgress")  # Check for egress rules
