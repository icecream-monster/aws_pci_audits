"""
This script is meant for a workaround because AWS SSO & Identity Store cannot generate a human-readable table for audits 
This script takes 1 command: the account id you want to scan
Modify the script so you can upload your: sso_arn, identity_store_id, and [profile] ahead of time
"""

import boto3
import boto3.session
import sys
from datetime import datetime
import openpyxl
from openpyxl import Workbook, load_workbook

aws_account_id = sys.argv[1]  # {aws_account_id}
your_sso_arn = "",  # {sso_arn}
is_id = ""  # {identity_store_id}

# Open workbook using openpyxl
workbook_name = datetime.now().strftime("%Y-%m-%d") + \
    "-SSO_permissionset_review.xlsx"
try:
    wb = load_workbook(filename=workbook_name)
except:
    wb = Workbook()
# New worksheet for aws_account_id
ws = wb.create_sheet(aws_account_id)
# Remove the annoying default named 'Sheet'
try:
    wb.remove(wb["Sheet"])
except:
    pass
# Worksheet headings
ws.append(["Group/User", "Type", "PrincipalId",
           "Role Name", "Permission Sets ARN"])

# Establish SDK session
my_session = boto3.session.Session(
    profile_name="",  # [profile]
    # Region doesn't matter here -  sso / iam are aws global services
    region_name="us-west-2"
)
sso = my_session.client('sso-admin')  # Using sso-admin SDK
ids = my_session.client('identitystore')  # Using sso-admin SDK

# Use a dict to hold every permisison_set (+ descriptions) per aws_account_id
my_ps_dict = {}

for item in sso.list_permission_sets_provisioned_to_account(  # Grab every permisison_set
        InstanceArn=your_sso_arn,
        AccountId=aws_account_id)["PermissionSets"]:
    describe_ps_dict = sso.describe_permission_set(  # Grab descriptions for every permisison_set
        InstanceArn=your_sso_arn,
        PermissionSetArn=item)['PermissionSet']
    # To get rid of the annoying datetime format in 'CreatedDate' key
    describe_ps_dict['CreatedDate'] = str(describe_ps_dict['CreatedDate'])

    my_ps_dict[item] = describe_ps_dict
# print(my_ps_dict)

# Use a new dict to hold all unique groups per aws_account_id, and attach all permission_sets applied to the group
my_psgroup_dict = {}

for ps in my_ps_dict.keys():
    for group in sso.list_account_assignments(  # Grab all groups for every permisison_set
            InstanceArn=your_sso_arn,
            AccountId=aws_account_id,
            PermissionSetArn=ps)["AccountAssignments"]:
        # print(group["PrincipalId"])

        # Putting stuff in ps_psgroup_dict
        if not (
                group["PrincipalId"] in my_psgroup_dict.keys()
        ):  # If the group_id is new, add as a key to ps_psgroup_dict, create an empty list to hold descriptions
            my_psgroup_dict[group["PrincipalId"]] = group
            my_psgroup_dict[group["PrincipalId"]]["PermissionSet"] = []
            my_psgroup_dict[group["PrincipalId"]]["PermissionSet"].append(
                my_ps_dict[ps])  # Add in the results found from ps dict
        else:  # For existing group_ids, append new permission_set to existing list
            my_psgroup_dict[group["PrincipalId"]]["PermissionSet"].append(
                my_ps_dict[ps])
# print(my_psgroup_dict)

for item in my_psgroup_dict:  # You either have GROUP or USER names
    if my_psgroup_dict[item]["PrincipalType"] == "GROUP":
        my_psgroup_dict[item].update({"Name": ids.describe_group(IdentityStoreId=is_id,
                                                                 GroupId=item)["DisplayName"]})

    else:  # User, no other options
        my_psgroup_dict[item].update({"Name": ids.describe_user(IdentityStoreId=is_id,
                                                                UserId=item)["UserName"]})

for item in my_psgroup_dict:
    for x in my_psgroup_dict[item]["PermissionSet"]:
        ws.append([my_psgroup_dict[item]["Name"],
                   my_psgroup_dict[item]["PrincipalType"], my_psgroup_dict[item]["PrincipalId"], ",".join([
                                                                                                          x["Name"]]),
                   ",".join([x["PermissionSetArn"]])])

wb.save(filename=workbook_name)
