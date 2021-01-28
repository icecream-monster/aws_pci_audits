"""
This script generates all PCI Security Groups per account, per region
This script takes 2 commands: aws_account_profile, and region
"""
import boto3
import boto3.session
import sys
from datetime import datetime
import openpyxl
from openpyxl import Workbook, load_workbook

aws_account_profile = sys.argv[1]
region = sys.argv[2]

my_session = boto3.session.Session(
    profile_name=aws_account_profile,
    region_name=region)  # Establshing a session with your profile
ec2 = my_session.client("ec2")  # Using ec2 sdk

# Open workbook using openpyxl
workbook_name = datetime.now().strftime("%Y-%m-%d") + "-pci_sgs.xlsx"
try:
    wb = load_workbook(filename=workbook_name)
except:
    wb = Workbook()
# New worksheet for region
ws = wb.create_sheet(region)
# Remove the annoying default named 'Sheet'
try:
    wb.remove(wb["Sheet"])
except:
    pass
# Worksheet headings
ws.append(["sg_id", "sg_name"])

# Get all ec2 instances with pci tags
response = ec2.describe_instances(
    Filters=[
        {
            "Name":
            "tag:Name",
            "Values": [
                # "Application", # For example, Name or Application, or Resource
                # "PCI", or whatever you've tagged your PCI resources with"
                # Chris Tan tagged all the PCI resources for me here. Thanks Chris!
            ]
        },
    ], )
# print(response["Reservations"][0]["Instances"][0]["NetworkInterfaces"][0]["Groups"])
for item in response["Reservations"][0]["Instances"][0]["NetworkInterfaces"][
        0]["Groups"]:
    sg_id = item["GroupId"]
    sg_name = item["GroupName"]
    ws.append([sg_id, sg_name])
wb.save(filename=workbook_name)