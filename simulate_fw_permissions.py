"""
This script is checks who has the permission to modify firewall rules: including both SGs and NACL on aws
This script takes 1 command: the account id you want to scan
NOTE: This script ONLY simulates USER roles that users use federation to sign on, aka., does not check for aws service roles
"""

import boto3
import boto3.session
import openpyxl
import sys 
from openpyxl import Workbook, load_workbook
from datetime import datetime
import json

aws_account_id = sys.argv[1]

my_session = boto3.session.Session(profile_name = aws_account_id) # Establshing a session with your profile 
iam = my_session.client('iam')

response = iam.list_roles(
    PathPrefix = "/aws-reserved/sso.amazonaws.com/us-west-2/" # Only user roles, not serivce roles 
)

# Open workbook using openpyxl
workbook_name = datetime.now().strftime("%Y-%m-%d") + "-simulate_fw_permissions.xlsx"
try:
    wb = load_workbook(filename=workbook_name)
except:
    wb = Workbook()
ws = wb.create_sheet(aws_account_id)
# Remove the annoying default named 'Sheet'
try:
    wb.remove(wb["Sheet"])
except:
    pass
# Worksheet headings
ws.append(["SSO Role", "SSO Arn", "Resources", "Action", "Simulate Policy Decision"])

# convert responses to json to better view 
# json_response = json.dumps(response)

sso_role_dict = {}

for item in response["Roles"]:
    item["CreateDate"] = str(item["CreateDate"]) # getting rid of the annoying date format
    sso_role_dict.update({item["RoleName"]:str(item["Arn"])})

# firewall permission to simulate
sg_permissions = ["ec2:AuthorizeSecurityGroupEgress", 
"ec2:AuthorizeSecurityGroupIngress", 
"ec2:CreateSecurityGroup", 
"ec2:DeleteSecurityGroup", 
"ec2:RevokeSecurityGroupEgress", 
"ec2:RevokeSecurityGroupIngress", 
"ec2:CreateNetworkAcl", 
"ec2:CreateNetworkAclEntry", 
"ec2:DeleteNetworkAcl", 
"ec2:DeleteNetworkAclEntry", 
"ec2:ModifyNetworkInterfaceAttribute", 
"ec2:ReplaceNetworkAclAssociation", 
"ec2:ReplaceNetworkAclEntry"]

for key, value in sso_role_dict.items():
    for perm in sg_permissions: 
        response = iam.simulate_principal_policy(
            PolicySourceArn=value,
            ActionNames=[ 
                perm
            ]
        )
        resource = response["EvaluationResults"][0]["EvalResourceName"]
        action = response["EvaluationResults"][0]["EvalActionName"]
        decision = response["EvaluationResults"][0]["EvalDecision"]

        ws.append([key, value, resource, action, decision])

wb.save(filename=workbook_name)
